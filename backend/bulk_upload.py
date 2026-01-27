"""
Bulk upload functionality for invoices and villas
"""
from fastapi import APIRouter, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from datetime import datetime, timedelta
import os
import logging
import uuid
import io
from typing import List, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

from auth import require_admin, require_manager_or_admin, require_accountant
from models import Invoice, Villa, INVOICE_TYPE_MAINTENANCE
from email_service import email_service

load_dotenv()

logger = logging.getLogger(__name__)

bulk_router = APIRouter(prefix="/bulk")

# Database connection
MONGO_URL = os.getenv('MONGO_URL', 'mongodb://localhost:27017')
DB_NAME = os.getenv('DB_NAME', 'test_database')


async def get_db():
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]
    return db, client


def generate_maintenance_invoice_number() -> str:
    """Generate unique invoice number for maintenance"""
    import random
    import string
    now = datetime.utcnow()
    random_suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=5))
    return f"TROA-MAINT-{now.year}{now.month:02d}-{random_suffix}"


# ============ INVOICE BULK UPLOAD ============

@bulk_router.get("/invoices/template")
async def download_invoice_template(request: Request):
    """Download sample Excel template for bulk invoice creation - Accountant, Manager, Admin"""
    await require_accountant(request)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Template"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Villa Number*",
        "Description*",
        "Quantity",
        "Rate*",
        "Discount Type",
        "Discount Value",
        "Due Days"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Sample data rows
    sample_data = [
        ["A-101", "Monthly Maintenance - January 2025", 1, 5000, "none", 0, 20],
        ["A-101", "Water Charges - January 2025", 1, 500, "", "", ""],
        ["A-102", "Monthly Maintenance - January 2025", 1, 5000, "percentage", 10, 30],
        ["A-103", "Monthly Maintenance - January 2025", 1, 5000, "fixed", 500, 15],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet(title="Instructions")
    instructions = [
        ["BULK INVOICE UPLOAD INSTRUCTIONS"],
        [""],
        ["Required Fields (marked with *):"],
        ["- Villa Number: Must match an existing villa in the system"],
        ["- Description: Description of the charge"],
        ["- Rate: Amount in INR"],
        [""],
        ["Optional Fields:"],
        ["- Quantity: Defaults to 1 if empty"],
        ["- Discount Type: 'none', 'percentage', or 'fixed' (defaults to 'none')"],
        ["- Discount Value: Percentage (0-100) or fixed amount in INR"],
        ["- Due Days: Number of days until payment is due (defaults to 20)"],
        [""],
        ["IMPORTANT NOTES:"],
        ["1. Multiple rows with the same Villa Number will be combined into ONE invoice"],
        ["2. Each row represents a line item in the invoice"],
        ["3. Discount is applied per invoice (not per line item)"],
        ["4. If a villa has multiple line items, only the FIRST row's discount/due days are used"],
        ["5. Emails will be sent to all registered email addresses for each villa"],
        [""],
        ["EXAMPLE:"],
        ["Villa A-101 has 2 rows -> Creates 1 invoice with 2 line items"],
        ["Villa A-102 has 1 row -> Creates 1 invoice with 1 line item"],
    ]
    
    for row_idx, row in enumerate(instructions, 1):
        cell = ws_instructions.cell(row=row_idx, column=1, value=row[0] if row else "")
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif row and row[0].startswith("IMPORTANT") or row[0].startswith("EXAMPLE"):
            cell.font = Font(bold=True, color="C00000")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws_instructions.column_dimensions['A'].width = 80
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=TROA_Invoice_Upload_Template.xlsx"}
    )


@bulk_router.post("/invoices/upload")
async def bulk_upload_invoices(request: Request, file: UploadFile = File(...)):
    """Bulk upload maintenance invoices from Excel file - Accountant, Manager, Admin"""
    user = await require_accountant(request)
    
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    
    db, client = await get_db()
    
    try:
        # Read file
        content = await file.read()
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        
        # Parse rows (skip header)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        if not rows:
            raise HTTPException(status_code=400, detail="No data found in file")
        
        # Group rows by villa number
        villa_data = {}
        for row_idx, row in enumerate(rows, 2):
            villa_number = str(row[0]).strip() if row[0] else None
            description = str(row[1]).strip() if row[1] else None
            quantity = float(row[2]) if row[2] else 1.0
            rate = float(row[3]) if row[3] else 0.0
            discount_type = str(row[4]).strip().lower() if row[4] else 'none'
            discount_value = float(row[5]) if row[5] else 0.0
            due_days = int(row[6]) if row[6] else 20
            
            if not villa_number:
                continue  # Skip empty rows
            
            if not description:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Row {row_idx}: Description is required"
                )
            
            if rate <= 0:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Row {row_idx}: Rate must be greater than 0"
                )
            
            if discount_type not in ['none', 'percentage', 'fixed', '']:
                discount_type = 'none'
            
            if villa_number not in villa_data:
                villa_data[villa_number] = {
                    'line_items': [],
                    'discount_type': discount_type if discount_type else 'none',
                    'discount_value': discount_value,
                    'due_days': due_days
                }
            
            villa_data[villa_number]['line_items'].append({
                'id': str(uuid.uuid4()),
                'description': description,
                'quantity': quantity,
                'rate': rate,
                'amount': quantity * rate
            })
        
        if not villa_data:
            raise HTTPException(status_code=400, detail="No valid data found in file")
        
        # Validate all villas exist
        villa_numbers = list(villa_data.keys())
        existing_villas = await db.villas.find(
            {"villa_number": {"$in": villa_numbers}},
            {"_id": 0}
        ).to_list(1000)
        existing_villa_map = {v['villa_number']: v for v in existing_villas}
        
        missing_villas = [vn for vn in villa_numbers if vn not in existing_villa_map]
        if missing_villas:
            raise HTTPException(
                status_code=400,
                detail=f"Villas not found in system: {', '.join(missing_villas)}"
            )
        
        # Create invoices
        created_invoices = []
        email_results = []
        
        for villa_number, data in villa_data.items():
            villa = existing_villa_map[villa_number]
            
            # Calculate totals
            subtotal = sum(item['amount'] for item in data['line_items'])
            
            discount_amount = 0.0
            if data['discount_type'] == 'percentage':
                discount_amount = subtotal * (data['discount_value'] / 100)
            elif data['discount_type'] == 'fixed':
                discount_amount = min(data['discount_value'], subtotal)
            
            total_amount = max(subtotal - discount_amount, 0)
            due_date = datetime.utcnow() + timedelta(days=data['due_days'])
            
            # Get primary email
            primary_email = villa.get('emails', [''])[0] if villa.get('emails') else ''
            
            invoice = Invoice(
                invoice_number=generate_maintenance_invoice_number(),
                invoice_type=INVOICE_TYPE_MAINTENANCE,
                villa_number=villa_number,
                user_email=primary_email,
                user_name='',
                maintenance_line_items=data['line_items'],
                subtotal=subtotal,
                discount_type=data['discount_type'],
                discount_value=data['discount_value'],
                discount_amount=discount_amount,
                total_amount=total_amount,
                due_date=due_date,
                created_by_email=user['email'],
                created_by_name=user.get('name', ''),
                audit_log=[{
                    'action': 'created',
                    'timestamp': datetime.utcnow().isoformat(),
                    'by_email': user['email'],
                    'by_name': user.get('name', ''),
                    'details': f"Bulk upload: Maintenance invoice created with {len(data['line_items'])} line items"
                }]
            )
            
            await db.invoices.insert_one(invoice.dict())
            created_invoices.append({
                'invoice_number': invoice.invoice_number,
                'villa_number': villa_number,
                'total_amount': total_amount,
                'line_items_count': len(data['line_items'])
            })
            
            logger.info(f"Bulk upload: Invoice {invoice.invoice_number} created for villa {villa_number}")
            
            # Send email notifications to all villa emails
            for email in villa.get('emails', []):
                try:
                    # Get user name if registered
                    villa_user = await db.users.find_one({"email": email}, {"_id": 0, "name": 1})
                    user_name = villa_user.get('name', '') if villa_user else ''
                    
                    await email_service.send_maintenance_invoice_raised(
                        recipient_email=email,
                        user_name=user_name,
                        invoice_number=invoice.invoice_number,
                        villa_number=villa_number,
                        total_amount=total_amount,
                        due_date=due_date.strftime("%d %b %Y"),
                        line_items=data['line_items']
                    )
                    email_results.append({'email': email, 'status': 'sent', 'villa': villa_number})
                except Exception as email_error:
                    logger.error(f"Failed to send invoice email to {email}: {email_error}")
                    email_results.append({'email': email, 'status': 'failed', 'villa': villa_number, 'error': str(email_error)})
        
        return {
            'success': True,
            'message': f"Successfully created {len(created_invoices)} invoice(s)",
            'invoices': created_invoices,
            'email_notifications': {
                'sent': len([e for e in email_results if e['status'] == 'sent']),
                'failed': len([e for e in email_results if e['status'] == 'failed']),
                'details': email_results
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing bulk invoice upload: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")
    finally:
        client.close()


# ============ VILLA BULK UPLOAD ============

@bulk_router.get("/villas/template")
async def download_villa_template(request: Request):
    """Download sample Excel template for bulk villa creation/update - Admin only"""
    await require_admin(request)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Villa Template"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Villa Number*",
        "Square Feet",
        "Emails (comma-separated)"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Sample data rows
    sample_data = [
        ["A-101", 2500, "owner1@email.com, co-owner@email.com"],
        ["A-102", 3000, "owner2@email.com"],
        ["B-201", 1800, "tenant@email.com, landlord@email.com"],
        ["C-301", 2200, ""],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet(title="Instructions")
    instructions = [
        ["BULK VILLA UPLOAD INSTRUCTIONS"],
        [""],
        ["Required Fields (marked with *):"],
        ["- Villa Number: Unique identifier for the villa (e.g., A-101, 205, B-302)"],
        [""],
        ["Optional Fields:"],
        ["- Square Feet: Area of the villa in square feet (defaults to 0)"],
        ["- Emails: Comma-separated list of email addresses associated with this villa"],
        [""],
        ["IMPORTANT NOTES:"],
        ["1. If a villa already exists, it will be UPDATED (upsert operation)"],
        ["2. For existing villas, new emails will be MERGED with existing ones"],
        ["3. Existing emails will NOT be removed - only new ones will be added"],
        ["4. Email addresses are automatically converted to lowercase"],
        ["5. Duplicate emails within a villa are automatically removed"],
        [""],
        ["EXAMPLE:"],
        ["If Villa A-101 exists with email 'old@email.com'"],
        ["And you upload with email 'new@email.com'"],
        ["Result: Villa A-101 will have both 'old@email.com' and 'new@email.com'"],
    ]
    
    for row_idx, row in enumerate(instructions, 1):
        cell = ws_instructions.cell(row=row_idx, column=1, value=row[0] if row else "")
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif row and (row[0].startswith("IMPORTANT") or row[0].startswith("EXAMPLE")):
            cell.font = Font(bold=True, color="059669")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50
    ws_instructions.column_dimensions['A'].width = 80
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=TROA_Villa_Upload_Template.xlsx"}
    )


@bulk_router.post("/villas/upload")
async def bulk_upload_villas(request: Request, file: UploadFile = File(...)):
    """Bulk upload/update villas from Excel file - Admin only"""
    user = await require_admin(request)
    
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    
    db, client = await get_db()
    
    try:
        # Read file
        content = await file.read()
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        
        # Parse rows (skip header)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        if not rows:
            raise HTTPException(status_code=400, detail="No data found in file")
        
        created_count = 0
        updated_count = 0
        results = []
        
        for row_idx, row in enumerate(rows, 2):
            villa_number = str(row[0]).strip() if row[0] else None
            square_feet = float(row[1]) if row[1] else 0.0
            emails_str = str(row[2]).strip() if row[2] else ''
            
            if not villa_number:
                continue  # Skip empty rows
            
            # Parse emails
            new_emails = []
            if emails_str:
                for email in emails_str.split(','):
                    email = email.strip().lower()
                    if email and '@' in email:
                        new_emails.append(email)
            
            # Check if villa exists
            existing_villa = await db.villas.find_one({"villa_number": villa_number}, {"_id": 0})
            
            if existing_villa:
                # Merge emails (keep existing, add new)
                existing_emails = [e.lower() for e in existing_villa.get('emails', [])]
                merged_emails = list(set(existing_emails + new_emails))
                
                # Update villa
                update_data = {
                    "updated_at": datetime.utcnow(),
                    "emails": merged_emails
                }
                
                # Only update square_feet if provided and non-zero
                if square_feet > 0:
                    update_data["square_feet"] = square_feet
                
                await db.villas.update_one(
                    {"villa_number": villa_number},
                    {"$set": update_data}
                )
                
                updated_count += 1
                results.append({
                    'villa_number': villa_number,
                    'action': 'updated',
                    'emails_added': len(set(new_emails) - set(existing_emails)),
                    'total_emails': len(merged_emails)
                })
                logger.info(f"Bulk upload: Villa {villa_number} updated by {user['email']}")
            else:
                # Create new villa
                villa = Villa(
                    villa_number=villa_number,
                    square_feet=square_feet,
                    emails=new_emails
                )
                
                await db.villas.insert_one(villa.dict())
                
                created_count += 1
                results.append({
                    'villa_number': villa_number,
                    'action': 'created',
                    'emails_added': len(new_emails),
                    'total_emails': len(new_emails)
                })
                logger.info(f"Bulk upload: Villa {villa_number} created by {user['email']}")
        
        return {
            'success': True,
            'message': f"Processed {len(results)} villa(s): {created_count} created, {updated_count} updated",
            'created': created_count,
            'updated': updated_count,
            'details': results
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing bulk villa upload: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")
    finally:
        client.close()
